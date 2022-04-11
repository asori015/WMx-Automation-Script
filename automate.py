import http.client
import sys
import openpyxl
import threading
import concurrent.futures
import argparse
import json
import gzip
import re
from datetime import datetime
import logging

all_columns = [
    "ORDER_CREATE_DATE_PST",
    "CASE_CREATE_DATE_PST",
    "MBOLKEY",
    "LOAD_ID",
    "TR_TYPE",
    "SITEID",
    "EXTERNKEY",
    "ORDERKEY",
    "CS_ID",
    "SSCC",
    "CONT_KEY",
    "MASTER_CONTAINERKEY",
    "CARRIER",
    "PICK_METHOD",
    "LANE",
    "ROUTE",
    "PACKGROUPKEY",
    "TOTALQTY",
    "COMMENTS"
]

WMxHeaderTable = {
	0	: ('Host', 'api.security.wmxp008.wmx.sc.xpo.com'),
	1	: ('Connection', 'keep-alive'),
	2	: ('Content-Length', '0'),
	3	: ('Accept', 'text/json'),
	4	: ('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'),
	5	: ('Content-Type', 'text/json'),
	6	: ('Origin', 'http://ont005.wmx.sc.xpo.com'),
	7	: ('Referer', 'http://ont005.wmx.sc.xpo.com/'),
	8	: ('Accept-Encoding', 'gzip, deflate'),
	9	: ('Accept-Language', 'en-US,en;q=0.9'),
	10	: ('Host', 'api.cqrs.wmxp008.wmx.sc.xpo.com'),
	11	: ('Xposc-EquipmentID', ''),
	12	: ('Xposc-DeviceID', 'PCAWS044'),
	13	: ('Authorization', ''),
    14	: ('Content-Type', 'application/json'),
	15	: ('Accept', 'application/json'),
	16	: ('Xposc-UserID', 'tponce'),
	17	: ('Xposc-SiteId', 'ONT005'),
	18	: ('Xposc-Language', 'ENG'),
	19	: ('Xposc-ClientID', '4044'),
	20	: ('Host', 'api.outbound.wmxp008.wmx.sc.xpo.com'),
    21	: ('Origin', 'http://m.ont005.wmx.sc.xpo.com'),
	22	: ('Referer', 'http://m.ont005.wmx.sc.xpo.com/'),
	24	: ('Xposc-EquipmentGroup', 'ALL'),
	25	: ('Xposc-ClientID', 'default'),
	26	: ('Host', 'api.datahelper.wmxp008.wmx.sc.xpo.com'),
    27	: ('Host', 'api.reporting.wmxp008.wmx.sc.xpo.com'),
}

BAxHeaderTable = {
    0	: ('Host', 'bax08s.am.gxo.com'),
	1	: ('Connection', 'keep-alive'),
	2	: ('sec-ch-ua', '" Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"'),
	3	: ('sec-ch-ua-mobile', '?0'),
	4	: ('sec-ch-ua-platform', '"Windows"'),
	5	: ('Upgrade-Insecure-Requests', '1'),
	6	: ('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36'),
	7	: ('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'),
	8	: ('Sec-Fetch-Site', 'none'),
	9	: ('Sec-Fetch-Mode', 'navigate'),
	10	: ('Sec-Fetch-User', '?1'),
	11	: ('Sec-Fetch-Dest', 'document'),
	12	: ('Accept-Encoding', 'gzip, deflate, br'),
	13	: ('Accept-Language', 'en-US,en;q=0.9'),
	14	: ('Origin', 'https://bax08s.am.gxo.com'),
	15	: ('Accept', 'text/css,*/*;q=0.1'),
	16	: ('Sec-Fetch-Site', 'same-origin'),
	29	: ('Sec-Fetch-Mode', 'no-cors'),
	30	: ('Sec-Fetch-Dest', 'style'),
	31	: ('Referer', 'https://bax08s.am.gxo.com/handm/login/'),
	32	: ('Cookie', 'session=eyJjc3JmX3Rva2VuIjoiN2U3NjIzNmJkYjU5YmNkMjc2YTExODQ4ZTc4YzhmMjg3N2RiYzBmYyJ9.Ykvx4w.ZZkcuqhmMP3m7QWqGoD3b-DR3u4; _client=handm; NSC_WTsw_l8t-johsftt.td.yqp.dpn443=ffffffffaf1b3d1045525d5f4f58455e445a4a42378b'),
	33	: ('Accept', '*/*'),
	34	: ('Sec-Fetch-Dest', 'script'),
	35	: ('Accept', 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8'),
	36	: ('Sec-Fetch-Dest', 'image'),
	37	: ('Content-Length', '0'),
	38	: ('Cache-Control', 'max-age=0'),
	39	: ('Content-Type', 'application/x-www-form-urlencoded'),
	40	: ('Cookie', '_client=handm; NSC_WTsw_l8t-johsftt.td.yqp.dpn443=ffffffffaf1b3d1045525d5f4f58455e445a4a42378b; session=.eJwljktuwzAMBe-idRbUj6RyGUOknlqjbQrYcTZF7143Wb4B3mB-gu_bXO7fH7iFaxAIp8w2rDbzkYR7jFoUoq4zqcgwp-nhEpa5YX8P1_t24BKOHdutf-F0fD7WtwPb2sMLL-s4aU5ZEuv_8bnBUlFTajh1SizZtbVJKihkVbjAoqVCfVSZRtKpWItEDgxI1C7V1aUxxwTqYLYCgmaNzTh28n4WVzZvM3OmEv3UWXGNhjLC7x-mZEhQ.Ykvy1Q.Qb1KS8rnORQsJhCTafzeX7nPxIA'),
	41	: ('Referer', 'https://bax08s.am.gxo.com/handm/superset/dashboard/1000000/'),
	42	: ('Cookie', '_client=handm; NSC_WTsw_l8t-johsftt.td.yqp.dpn443=ffffffffaf1b3d1045525d5f4f58455e445a4a42378b; session=.eJwlj0tuwzAMRO-idRbUj6RyGYOkqNaomwB23E3Ru1dJljPAPLz5DXbsY3ncv_wWroGcMGXUrrWp9UQoMXJhJzYeiYm6GgwLl7CM3Y_PcH3sp1_Cefh-k2-fjO1n_Th9XyW862Xts80pU0J-Dl_ZkarXlJpPHANSNm5tAJMX0EpYXKOmAtIrDQUSKNoigLl3p8hC1dioIcbkII6oxcE5c2yKUcBkGldUayNjhhJt4rQYR_XSp8h2N9mexvP63z-4X0yB.Ykvy2g.PrThQrvdHDybV1FCWXeaVjjGDow'),
	43	: ('X-CSRFToken', 'IjdlNzYyMzZiZGI1OWJjZDI3NmExMTg0OGU3OGM4ZjI4NzdkYmMwZmMi.Ykvy2g.yIkUUMhCbgy9ltl921oYWEFBLyM'),
	44	: ('Sec-Fetch-Mode', 'same-origin'),
	45	: ('Sec-Fetch-Dest', 'empty'),
	46	: ('Content-Type', 'multipart/form-data; boundary=----WebKitFormBoundary7XXhQdTyIALZLRcX'),
	48	: ('Content-Type', 'multipart/form-data; boundary=----WebKitFormBoundaryeoW7XBYpazf0Y0uG'),
	49	: ('Content-Type', 'multipart/form-data; boundary=----WebKitFormBoundary9nA69TfFQg8qZOib'),
	50	: ('Content-Type', 'multipart/form-data; boundary=----WebKitFormBoundaryRLLy0TetMfhPMq50')
}

def addThrees(value):
    newValue = ''
    for row in value:
        newValue += '3' + row
    return newValue

def getHeaders(indexes: list, headerTable: dict) -> dict:
    headerDict = {}
    for row in indexes:
        headerDict.update({headerTable[row][0]: headerTable[row][1]})
    return headerDict

def makeRequest(conn, method, path, headers: dict, payload: str):
	conn.request(method, path, payload, headers)
	res = conn.getresponse()
	print(str(res.status) + ', ' + str(res.reason))
	data = res.read()
	try:
		return data.decode('utf-8'), res.headers
	except:
		return gzip.decompress(data).decode('utf-8'), res.headers

def requestBAx(username: str, password: str) -> list:
    import functools
    conn = http.client.HTTPSConnection("bax08s.am.gxo.com", 443)
    logging.info('Connected to BAx')

    headers = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    response, responseHeaders = makeRequest(conn, "GET", "/handm/login/", getHeaders(headers, BAxHeaderTable), None)
    csrfToken = re.findall(r'csrf.*value="(.*)">', response)[0]
    cookieHeaders = re.findall(r'Set-Cookie: (.*);', str(responseHeaders))
    cookieHeader = functools.reduce(lambda a, b: a + b[:b.find(';')] + '; ', cookieHeaders, '')[:-2]

    headers = [0, 1, 37, 38, 2, 3, 4, 5, 14, 39, 6, 7, 16, 9, 10, 11, 31, 12, 13, 32]
    params = 'csrf_token=' + csrfToken + '&version=1.2.4&username=' + username + '&password=' + password + '&execution=&_eventId=submit&geolocation=&submit=Log+In'
    BAxHeaderTable[37] = ('Content-Length', str(len(params)))
    BAxHeaderTable[32] = ('Cookie', cookieHeader)
    
    logging.info('Logging into to BAx...')
    makeRequest(conn, "POST", "/handm/login/", getHeaders(headers, BAxHeaderTable), params)

    logging.info('Loading main dashboard...')
    headers = [0, 1, 38, 5, 6, 7, 16, 9, 10, 11, 2, 3, 4, 31, 12, 13, 40]
    makeRequest(conn, "GET", "/handm/superset/dashboard/1000000/", getHeaders(headers, BAxHeaderTable), None)
    
    logging.info('Loading random API...')
    headers = [0, 1, 2, 43, 3, 6, 4, 33, 16, 44, 45, 41, 12, 13, 42]
    makeRequest(conn, "GET", "/csstemplateasyncmodelview/api/read", getHeaders(headers, BAxHeaderTable), None)

    logging.info('Loading another random API...')
    headers = [0, 1, 2, 43, 3, 6, 4, 33, 16, 44, 45, 41, 12, 13, 42]
    makeRequest(conn, "GET", "/handm/csstemplateasyncmodelview/api/read", getHeaders(headers, BAxHeaderTable), None)

    logging.info('Loading the Aging Totes Report...')
    headers = [0, 1, 37, 2, 43, 3, 49, 6, 4, 33, 14, 16, 44, 45, 41, 12, 13, 42]
    params = '{"datasource":"1004028__table","viz_type":"table","slice_id":1002187,"granularity_sqla":null,"time_grain_sqla":"P1D","time_range":"No filter","groupby":[],"metrics":[],"percent_metrics":[],"timeseries_limit_metric":null,"row_limit":10000,"include_time":false,"order_desc":true,"all_columns":["ORDER_CREATE_DATE_PST","CASE_CREATE_DATE_PST","MBOLKEY","LOAD_ID","TR_TYPE","SITEID","EXTERNKEY","ORDERKEY","CS_ID","SSCC","CONT_KEY","MASTER_CONTAINERKEY","CARRIER","PICK_METHOD","LANE","ROUTE","PACKGROUPKEY","TOTALQTY","COMMENTS"],"order_by_cols":[],"adhoc_filters":[],"table_timestamp_format":"%Y-%m-%d","page_length":0,"include_search":false,"table_filter":false,"align_pn":false,"color_pn":true,"label_colors":{},"extra_filters":[]}'
    BAxHeaderTable[61] = ('Content-Length', str(len(params)))
    response, responseHeaders = makeRequest(conn, "POST", "/handm/superset/explore_json/?form_data=%7B%22slice_id%22%3A1002187%7D", getHeaders(headers, BAxHeaderTable), params)
    atr = json.loads(response)
    return atr['data']['records']

def initWMx() -> http.client.HTTPConnection:
    print('Initializing WMx connection')
    conn = http.client.HTTPConnection("172.19.45.163", 80)
    
    headers = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    params = '{"UserId":"tponce","Psw":"Welcome1234567","SiteId":"ONT005"}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    response, responseHeaders = makeRequest(conn, "POST", "/login", getHeaders(headers, WMxHeaderTable), params)
    WMxHeaderTable[13] = ('Authorization', 'Bearer ' + json.loads(response)['Token'])
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/equipmentgroupdd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/clientdd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/deviceprinters/5043415753303434", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/userprofile/74706f6e6365", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/listboxdd/434c49454e54", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/usersearch/57585f434f4e5441494e45525f53484950/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '{"CONTAINERKEY":"0002401084"}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/queryservice/data/57585f434f4e5441494e45525f53484950/4144445453/66616c7365/31/3330", getHeaders(headers, WMxHeaderTable), params)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/searchconfig/57585f434f4e5441494e45525f53484950", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/listboxdd/434f4e5441494e4552535441545553", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containerstatusdd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/cartonizecodesdd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containership/", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/434f4e5441494e455244454641554c54484549474854", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/434e54424c4443415054555245494e445543544c4f43", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/usersearch/57585f434f4e5441494e45525f5348495044544c/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/searchconfig/57585f434f4e5441494e45525f5348495044544c", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/554944544c5041474553495a45", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/searchconfig/57585f434f4e5441494e45525f5348495044544c", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)

    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)
    return conn

def getNewLoadID(conn):
    print('looking for new Dummy Load ID...')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/4d4f42494c45545241494c45524c4f4144414456534541524348", getHeaders(headers, WMxHeaderTable), None)
    # Here we get all relevant load IDs and look for the first B05 load that's status 101 New
    headers = [20, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/container/activeloads", getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/searchconfig/57585f4c4f4144", getHeaders(headers, WMxHeaderTable), None)
    #
    activeLoads = json.loads(response)
    dummyLoads = []
    for load in activeLoads:
        if load['TRAILERTYPE'] == 'B05' and load['STATUS'] == 101:
            dummyLoads.append(load['LOADID'])
    #
    if len(dummyLoads) == 0:
        print('No B05 Load IDs found. What the hell happened?')
        input('Press anything to continue...')
        return ''
    dummyLoads.sort()
    dummyLoad = dummyLoads[0]
    print('Found LOAD_ID: ' + dummyLoad)
    if input('Use this Load ID? (Y/N)') != 'Y':
        return ''
    print('Utilizing Dummy Load: ' + dummyLoad)
    #
    headers = [10, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '{"LOADID":"' + dummyLoad + '"}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/queryservice/data/57585f4c4f4144/4144445453/66616c7365/31/3330", getHeaders(headers, WMxHeaderTable), params)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/load/" + addThrees(dummyLoad), getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/usersearch/57585f4c4f414444544c/4c4f41444d414e4147454d454e54", getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/4c4f414444544c42594f52444552", getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/loadDtlList/" + addThrees(dummyLoad), getHeaders(headers, WMxHeaderTable), None)

    # Updating our chosen load ID with correct values
    response = json.loads(response)
    response['SEALNUMBER'] = 'SN' + dummyLoad
    response['TRAILERNUMBER'] = 'DUMMY' + date.today().strftime('%m%d%y')
    response['DOOR'] = 'S9'
    response['LOADTYPE'] = 'VIA_SP'
    response = json.dumps(response)
    #
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = response
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/isloaddirty", getHeaders(headers, WMxHeaderTable), params)
    #
    headers = [26, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = response
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/datamodify/saveload", getHeaders(headers, WMxHeaderTable), params)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/load/" + addThrees(dummyLoad), getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/4c4f414444544c42594f52444552", getHeaders(headers, WMxHeaderTable), None)
    #
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/loadDtlList/" + addThrees(dummyLoad), getHeaders(headers, WMxHeaderTable), None)
    return dummyLoad

def handle_118(conn):
    print('TODO: 118')
    return None
    # print('118\'s are handled in')
    # This request just fails...
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/usersearch/57585f57415645/57415645", headers, None)
    # Here we're grabbing every wave that's status 118 or below.
    headers = [10, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '{"STATUS":"102|105|106|112|115|116|118|101"}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    response, responseHeaders = makeRequest(conn, "PUT", "/queryservice/data/57585f57415645/4144445453/66616c7365/31/3330", getHeaders(headers, WMxHeaderTable), params)
    #
    response = json.loads(response)
    numberOfWaves = response['Count']
    currentWaves = response['Data']
    pageNumber = 2
    while numberOfWaves > 30:
        headers = [10, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        params = '{"STATUS":"102|105|106|112|115|116|118|101"}'
        WMxHeaderTable[2] = ('Content-Length', str(len(params)))
        response, responseHeaders = makeRequest(conn, "PUT", "/queryservice/data/57585f57415645/4144445453/66616c7365/" + addThrees(str(pageNumber)) + "/3330", getHeaders(headers, WMxHeaderTable), params)
        #
        response = json.loads(response)
        currentWaves += response['Data']
        numberOfWaves -= 30
        pageNumber += 1
    #
    casesInWaves = []
    for wave in currentWaves:
        if wave['LINECOUNT'] == 0:
            continue
        waveKey = wave['WAVEKEY']
        print(waveKey)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/wavecodefilteroperatorsdd", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/wavestatusdd", getHeaders(headers, WMxHeaderTable), None)
        #
        # headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        # params = '{}'
        # lookupHeaders[2] = ('Content-Length', str(len(params)))
        # response, responseHeaders = makeRequest(conn, "PUT", "/waveprocess/updwavestatus/" + addThrees(waveKey), headers, params)
        # response = json.loads(response)
        # if response['LINECOUNT'] == 0:
        #     continue
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/wavedetailsbywavekey/" + addThrees(waveKey), getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/usersearch/434f52454f52445f5741564544544c5f5657/57415645", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/searchconfig/434f52454f52445f5741564544544c5f5657", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/usersearch/57585f4f524445525f43415345/57415645", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/searchconfig/57585f4f524445525f43415345", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/searchconfig/57585f4f524445525f43415345", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/listboxdd/4f52444552535441545553", getHeaders(headers, WMxHeaderTable), None)
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        response, responseHeaders = makeRequest(conn, "GET", "/queryservice/ordercasebywavekey/" + addThrees(waveKey), getHeaders(headers, WMxHeaderTable), None)
        # print(response)
        if len(response) > 2:
            response = json.loads(response)
            casesInWaves += response
        #
        headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
        makeRequest(conn, "GET", "/queryservice/autoprocesswave/" + addThrees(waveKey), getHeaders(headers, WMxHeaderTable), None)
        #
    #
    print(casesInWaves)
    
def handle_125():
    print('125\'s are handled in Grey Orange. Skipping...')
    return None

def handle_130():
    print('130\'s are handled in Grey Orange. Skipping...')
    return None

def handle_135(conn, sscc):
    print('Handling status 135 tote')
    response = ''
    caseID = ''
    containerKey = ''

    print('1')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/order/" +  addThrees(sscc) + "/case", getHeaders(headers, WMxHeaderTable), None)
    caseID = re.findall(r'"CASEID":"(\d*)"', response)[0]
    
    print('2')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = ''
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    response, responseHeaders = makeRequest(conn, "PUT", "/container/opencontainersingle/" + addThrees(caseID) + "/", getHeaders(headers, WMxHeaderTable), params)
    containerKey = response[1:-1]
    
    print('3')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containership/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    
    print('5')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '{}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/addcasesingle/" + addThrees(containerKey) + "/" + addThrees(caseID) + "/", getHeaders(headers, WMxHeaderTable), params)
    
    print('6')
    #The response from this query gets fed into the 'iscontainerdirty' request
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/containership/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    print('7')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/carrierdd", getHeaders(headers, WMxHeaderTable), None)
    print('8')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containershipdetailsbycontainerkey/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    print('9')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/434e545041434b4147455459504557495448434152544f4e54595045", getHeaders(headers, WMxHeaderTable), None)
    print('10')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = response
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/iscontainerdirty", getHeaders(headers, WMxHeaderTable), params)
    print('11: Requesting to close container')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '{}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/closepackedcontainersingle/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), params)
    print('12')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containership/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    print('13')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/containershipdetailsbycontainerkey/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    print('14')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/434e545041434b4147455459504557495448434152544f4e54595045", getHeaders(headers, WMxHeaderTable), None)
    return containerKey

def handle_140(conn, sscc):
    print('Handling status 140 tote')
    return handle_135(conn, sscc) # Status 140 totes function almost the same as status 135 totes

def handle_141(conn, sscc):
    print('TODO: 141')
    return None
    print('Handling status 141 tote')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/orderstatusdd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/verdoctypedd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/config/434f5245564552494659/5645524849444544455441494c", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/564552494649434154494f4e534541524348", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/4f52444552564552494659", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configpropattr/434f4e5441494e455253494e474c45", getHeaders(headers, WMxHeaderTable), None)
    # Entering in an SSCC value
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/order/" + addThrees(sscc) + "/case", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = '""'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/verification/create/43415345/" + addThrees(sscc), getHeaders(headers, WMxHeaderTable), params)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/order/" + addThrees(sscc) + "/case", getHeaders(headers, WMxHeaderTable), None)
    response = json.loads(response)
    caseID = response['CASEID']
    orderKey = response['ORDERKEY']
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/usersearch/434f52455645525f4f524445525f5645524946595f5657/564552494649434154494f4e", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/searchconfig/434f52455645525f4f524445525f5645524946595f5657", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/cartonizecodedd", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/tracebydoctypedockey/open/4f52444552/" + addThrees(orderKey), getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/listboxdd/4f52444552535441545553", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/listboxdd/4f52444552535441545553", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/tracebydoctypedockey/open/43415345/" + addThrees(caseID), getHeaders(headers, WMxHeaderTable), None)
    
    headers = [20, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/verification/gervergrpdd/43415345/" + addThrees(sscc), getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/order/" + addThrees(sscc) + "/case", getHeaders(headers, WMxHeaderTable), None)
    
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    params = response.decode('utf-8')
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/verification/instructions/case/43415345", getHeaders(headers, WMxHeaderTable), params)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    makeRequest(conn, "GET", "/queryservice/order/" + addThrees(orderKey), getHeaders(headers, WMxHeaderTable), None)
    
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/orderverifyviewbydoctype/43415345/" + addThrees(caseID), getHeaders(headers, WMxHeaderTable), None)
    response = json.loads(response)
    
    for skuRecord in response:
        count = skuRecord['PICKQTY'] - skuRecord['VERIFYQTY']
        while count > 0:
            headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            makeRequest(conn, "GET", "/queryservice/codelist/56455251545945444954", getHeaders(headers, WMxHeaderTable), None)
            
            headers = [27, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            params = '"' + skuRecord['SKU'] + '"'
            WMxHeaderTable[2] = ('Content-Length', str(len(params)))
            makeRequest(conn, "POST", "/format/parselist/534b554c4142454c/3030/3138", getHeaders(headers, WMxHeaderTable), params)
            
            headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            makeRequest(conn, "GET", "/queryservice/tracebydoctypedockey/open/534b55/" + addThrees(sku), getHeaders(headers, WMxHeaderTable), None)
            
            headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            makeRequest(conn, "GET", "/queryservice/tracebydoctypedockey/open/564552534b55/" + addThrees(sku), getHeaders(headers, WMxHeaderTable), None)
            
            headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            params = '{"ORDERVERIFYID":27110881,"SITEID":"ONT005","CASEID":"0011847104","ORDERKEY":"0003772380","ORDERLINENO":9,"SKU":"000990977002214005","DESCRIPTION":"Sweater White, L","SERIAL_FLAG":"N","CLIENTID":"4044","LOT":"LOT","PICKQTY":3,"VERIFYQTY":0,"UOM":"EA","UOM_LEVEL":1,"UOMQTY":3,"UOMCONVQTY":1,"NONINVENTORY_FLAG":"N","DATACAPTURE_FLAG":"N","DATACAPTURECODE":null,"VALIDATE_FLAG":"N","VALIDATECODE":null,"VERIFYWHO":null,"DEFECT_FLAG":"N","DEFECTCODE":null,"DEFECTNOTES":null,"PICKER":null,"RESOLVECODE":null,"RESOLVEWHO":null,"HAZMAT_FLAG":"N","HAZMATCODE":null,"PACK_FLAG":"N","PACKTS":null,"PACKWHO":null,"PACKREFKEY":null,"DROPID":null,"ORDERPICKID":33851009,"STATUS":141,"STATUSTS":"2022-03-12T18:16:07.180676-05:00","VERGROUPCD":null,"VERGROUPKEY":null,"ADDTS":"2022-03-12T18:16:07.180682-05:00","ADDWHO":"yharosvargas","EDITTS":"2022-03-12T18:16:07.180685-05:00","EDITWHO":"yharosvargas","SKUSCAN_FLAG":"Y","LOTATR1":null,"LOTATR2":null,"LOTATR3":null,"LOTATR4":null,"LOTATR5":null,"LOTATR6":null,"LOTATR7":null,"LOTATR8":null,"CARTONTYPE":"A16"}'
            WMxHeaderTable[2] = ('Content-Length', str(len(params)))
            makeRequest(conn, "PUT", "/verification/instructions/verifyline", getHeaders(headers, WMxHeaderTable), params)
            
            headers = [20, 1, 2, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            params = '""'
            WMxHeaderTable[2] = ('Content-Length', str(len(params)))
            makeRequest(conn, "PUT", "/verification/verifydetail/3237313130383831/31", getHeaders(headers, WMxHeaderTable), params)
            
            headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            makeRequest(conn, "GET", "/queryservice/sku/" + addThrees(sku), getHeaders(headers, WMxHeaderTable), None)
            
            headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 6, 7, 8, 9]
            makeRequest(conn, "GET", "/queryservice/orderverifyviewbydoctype/43415345/" + addThrees(caseID), getHeaders(headers, WMxHeaderTable), None)

def handle_150(conn, sscc):
    print('Handling status 150 tote')
    return handle_135(conn, sscc) # Status 150 totes function almost the same as status 135 totes

def handle_160():
    print('Handling status 160 tote') # We don't really need to do anything for this status.

def handle_161(conn, sscc, row):
    print('Handling status 161 tote')
    unmasteredTotes[row[11].value] = row[10].value

def handle_165(row):
    print('Handling status 165 tote')
    masteredTotes[row[11].value] = row[10].value

def masterTotes(conn, containerKey):
    print('Mastering stack...')
    print('1: Looking up master container')
    headers = [20, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/container/lookupmastercontainer/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), None)
    
    print('2: Closing master container')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    params = '{}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/closemastercontainer/" + addThrees(containerKey), getHeaders(headers, WMxHeaderTable), params)
    return containerKey

def initLoading(conn, loadID):
    print('Loading all totes to Load ID:' + loadID)
    
    #First we enter the loadID
    print('1')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/loadSearch/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)

    print('2')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    response, responseHeaders = makeRequest(conn, "GET", "/queryservice/load/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)
    
    print('3')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/uiconfig/4d4f42494c454c4f4144/4d4f42494c45", getHeaders(headers, WMxHeaderTable), None)
    
    print('4')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/trailertypedd", getHeaders(headers, WMxHeaderTable), None)
    
    print('5')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/uiconfig/4d4f42494c454c4f4144/4d4f42494c45", getHeaders(headers, WMxHeaderTable), None)
    
    print('6')
    headers = [10, 1, 11, 12, 13, 14, 15, 16, 4, 17, 18, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/uiconfig/4d4f42494c454c4f4144/4d4f42494c45", getHeaders(headers, WMxHeaderTable), None)
    #Then we process the load ID
    
    print('7')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    params = response.decode('utf-8')
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/isloaddirty", getHeaders(headers, WMxHeaderTable), params)
    
    print('8')
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    params = response.decode('utf-8')
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/saveload", getHeaders(headers, WMxHeaderTable), params)
    
    print('9')
    headers = [20, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/container/activeloads", getHeaders(headers, WMxHeaderTable), None)
    
    print('10')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/configbyconfigid/all/4c4f4144444f4f52434f4e4649524d", getHeaders(headers, WMxHeaderTable), None)
    
    print('11')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/load/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)
    
    print('12')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/loadDtlList/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)

def loadTotes(conn, loadID, containerKey):
    #This is us loading a containerKey
    print('13: Loading container ' + containerKey)
    headers = [20, 1, 2, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    params = '{}'
    WMxHeaderTable[2] = ('Content-Length', str(len(params)))
    makeRequest(conn, "PUT", "/container/loadcontainer/" + addThrees(loadID) + "/" + addThrees(containerKey) + "/5339/302e646f76716f386a6c366774", getHeaders(headers, WMxHeaderTable), params)
    
    print('14')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/load/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)
    
    print('15')
    headers = [10, 1, 11, 12, 13, 14, 15, 18, 24, 17, 16, 4, 19, 21, 22, 8, 9]
    makeRequest(conn, "GET", "/queryservice/loadDtlList/" + addThrees(loadID), getHeaders(headers, WMxHeaderTable), None)

def handleTote(conn: http.client.HTTPConnection, record: dict) -> str:
    status = record['record']['COMMENTS'][:3]

    if status == '118':
        return handle_118()
    elif status == '125':
        return handle_125()
    elif status == '130':
        return handle_130()
    elif status == '135':
        try:
            return handle_135(conn, record['record']['CS_ID'])
        except Exception as e:
            print('Error Occcured: 135')
            logging.exception(e)
    elif status == '140':
        try:
            return handle_140(conn, record['record']['CS_ID'])
        except Exception as e:
            print('Error Occcured: 140')
            logging.exception(e)
    elif status == '141':
        try:
            return handle_141(conn, record['record']['CS_ID'])
        except Exception as e:
            print('Error Occcured: 141')
            logging.exception(e)
    elif status == '150':
        try:
            return handle_150(conn, record['record']['CS_ID'])
        except Exception as e:
            print('Error Occcured: 150')
            logging.exception(e)
    elif status == '161':
        try:
            return None
            # masterTotes()
            # return handle_150(conn, record['record']['CS_ID'])
        except Exception as e:
            print('Error Occcured: 161')
            logging.exception(e)
    else:
        if status != 'COM':
            print('Tote status ' + status + ' is not handled by this script.')
    return None

def run(threadID: int, logLock: threading.Lock, args: argparse.Namespace, processedRecords: dict) -> None:
    # Apparently, threaded functions don't display exceptions normally, so
    # I'm encapsulating the entire function in a try-catch block to fix that.
    try:
        conn = initWMx()
        for index in range(len(processedRecords['unprocessedTotes'])):
            record = None
            with logLock:
                record = processedRecords['unprocessedTotes'][index]
                if record['threadid'] == 0:
                    record['threadid'] = threadID

            if record['threadid'] == threadID:
                containerKey = handleTote(conn, record)
                if containerKey != None:
                    with logLock:
                        record['threadid'] = 0
                        record['record']['CONT_KEY'] = containerKey
                        processedRecords['unloadedTotes'].append(record)

        initLoading(conn)
        for index in range(len(processedRecords['unloadedTotes'])):
            with logLock:
                record = processedRecords['unloadedTotes'][index]
                if record['threadid'] == 0:
                    record['threadid'] = threadID

            if record['threadid'] == threadID:
                loadTotes(conn, args.loadid, record['record']['CONT_KEY'])
    except Exception as e:
        logging.exception(e)
    return

def processATR(atr: list, blacklist: list) -> dict:
    unprocessedTotes = []
    unloadedTotes = []
    unmasteredTotes = {} # Dictionary for storing only *one* container for each masterbuild
    masteredTotes = {} # Same thing but for 161's


    unprocessedStatuses = ['118', '120', '125', '130', '135', '140', '141', '150']

    for record in atr:
        status = record['COMMENTS'][:3]
        if record['PACKGROUPKEY'][3:6] in blacklist:
            continue
        if record['CASE_CREATE_DATE_PST'][:10] > datetime.now().strftime('%m/%d/%y'):
            logging.debug('Skipping record: %s %s', record)
            continue
        if status in unprocessedStatuses:
            unprocessedTotes.append({'record': record, 'threadid': 0})
        elif status == '160':
            unloadedTotes.append({'record': record, 'threadid': 0})
        elif status == '161':
            unmasteredTotes[record['MASTER_CONTAINERKEY']] = record
        elif status == '165':
            masteredTotes[record['MASTER_CONTAINERKEY']] = record
    
    for masterContainerKey in unmasteredTotes:
        unprocessedTotes.append({'record': unmasteredTotes[masterContainerKey], 'threadid': 0})
    for masterContainerKey in masteredTotes:
        unloadedTotes.append({'record': masteredTotes[masterContainerKey], 'threadid': 0})

    return {'unprocessedTotes': unprocessedTotes, 'unloadedTotes': unloadedTotes}

def loadExcel(workBook: openpyxl.Workbook) -> list:
    records = []
    try:
        currentSheet = workBook['Sheet1']
    except:
        logging.exception('"Sheet1" not found, shutting down script process.')
        sys.exit()
    for index1, row in enumerate(currentSheet):
        if index1 == 0:
            continue
        record = {}
        for index2, cell in enumerate(row):
            record[all_columns[index2]] = cell.value
        records.append(record)
    return records

def dumpExcel(atr: list) -> None:
    from openpyxl.styles import PatternFill, Border, Side, Font

    wb = openpyxl.Workbook()
    newSheets = {}
    colunmWidths = [25.29, 23.29, 29.14, 11.29, 8.14, 7.57, 11.57, 11.29, 11.29, 22.57, 11.29, 23.29, 12.29, 13.86, 5.43, 6.57, 15.14, 9.86, 255]

    # Creating color and border styles
    red = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    yellow = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
    green = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050')
    thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Iterate through ATR and assign each record to correct Excel sheet
    for i, record in enumerate(atr):
        splitpoint = record['CARRIER']

        # Assign color and note to record
        color = None
        if record['COMMENTS'][:3] < '135':
            color = red
            if record['MBOLKEY'] == None: record['MBOLKEY'] = 'NEEDS IT SUPPORT'
        elif record['COMMENTS'][:3] < '180':
            color = yellow
            if record['MBOLKEY'] == None: record['MBOLKEY'] = 'PROCESSED, ASSUMED SHIPPED'
        else:
            color = green
            if record['MBOLKEY'] == None: record['MBOLKEY'] = 'ON DOCK'

        # Make new Excel sheet for each splitpoint that exists
        if splitpoint not in newSheets:
            newSheets[splitpoint] = {'sheet': wb.create_sheet(splitpoint[2:5] + ' ' + datetime.now().strftime('%m.%d.%y')), 'index': 0}
            for j, name in enumerate(all_columns):
                cellName = chr(j + ord('A')) + '1'
                newSheets[splitpoint]['sheet'][cellName].value = name
                newSheets[splitpoint]['sheet'][cellName].border = thinBorder
                newSheets[splitpoint]['sheet'][cellName].font = Font(b=True)
            newSheets[splitpoint]['sheet']['C1'] = 'NOTES'

        # Add record to correct Excel sheet
        for j, key in enumerate(record):
            cellName = chr(j + ord('A')) + str(newSheets[splitpoint]['index'] + 2)
            newSheets[splitpoint]['sheet'][cellName].value = record[key]
            newSheets[splitpoint]['sheet'][cellName].fill = color
        newSheets[splitpoint]['index'] += 1
    for sheet in wb:
        for i in range(len(all_columns)):
            sheet.column_dimensions[chr(ord('A') + i)].width = colunmWidths[i]

    wb.save(filename = 'Resources/OTR ' + datetime.now().strftime('%m.%d.%y %H.%M.%S') + '.xlsx')
    return

def getAgingToteReport(args: argparse.Namespace) -> list:
    wb = None
    atr = None

    # Load ATR
    if args.excelfile != None:
        # Check to see if the filename is valid
        try:
            with open(args.excelfile, 'r') as f:
                pass
            wb = openpyxl.load_workbook(filename = args.excelfile)
            logging.info('Opened %s', args.excelfile)
        except FileNotFoundError as e:
            logging.exception('File name not valid')
            logging.info('Falling back on BAx')
            atr = requestBAx(args.username, args.password)
        except Exception as e:
            logging.exception('File is not valid Excel file')
            logging.info('Falling back on BAx')
            atr = requestBAx(args.username, args.password)
    else:
        logging.info('No file found, using BAx')
        atr = requestBAx(args.username, args.password)

    if wb != None:
        atr = loadExcel(wb)
    return atr

def init_argparse() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        usage="%(prog)s [OPTION] [FILE]...",
        description="Automates WMx Dummy Load",
        epilog="Please use responsibly..."
    )
    parser.add_argument(
        "-v", "--version", 
        action="version",
        version = f"{parser.prog} version 1.0.0"
    )
    parser.add_argument(
        "-u", "--username",
        type=str,
        required=True,
        help='WMx username to login.'
    )
    parser.add_argument(
        "-p", "--password",
        type=str,
        required=True,
        help='WMx password to login.'
    )
    parser.add_argument(
        '-l', '--loadid',
        type=str,
        help='Load ID to be used for loading.'
    )
    parser.add_argument(
        '-t', '--threads',
        type=int,
        help='Number of threads this script will generate.'
    )
    parser.add_argument(
        "-e", "--excelfile",
        type=str,
        help='Excel file containing Aging Totes Report to be processed and loaded to dummy load. If left empty, script will directly grab ATR from Bax.'
    )
    parser.add_argument(
        "-f", "--format",
        action = 'store_true',
        help='Format BAx Aging Totes Report into Excel sheet.'
    )
    return parser

def main() -> None:
    # Set up logging for this script
    logging.basicConfig(format='(%(name)s - %(levelname)s) %(message)s', level=logging.DEBUG, filename='log.txt', filemode='w')

    # Parse CLI arguments
    parser = init_argparse()
    args = parser.parse_args()
    logging.debug('args = %s', args)

    # Get ATR
    atr = getAgingToteReport(args)
    logging.debug(atr)

    storeBlacklist = ['108', '196'] # Stores that will not be worked on by the script
    processedRecords = processATR(atr, storeBlacklist)
    logging.debug(processedRecords)

    if args.format:
        logging.info('Generating Open Tote Report Excel sheet')
        dumpExcel(atr)

    dummyLoad = args.loadid
    if dummyLoad == None or len(dummyLoad) != 10:
        conn = initWMx()
        dummyLoad = getNewLoadID(conn)
        conn.close()

    # Create threads
    logLock = threading.Lock()
    logging.info('Processing Aging Tote Report with %s threads', args.threads)
    if args.threads is not None and args.threads > 0:
        with concurrent.futures.ThreadPoolExecutor(max_workers=args.threads) as executor:
            executor.map(lambda x: run(x + 1, logLock, args, processedRecords), range(args.threads))

    logging.debug(processedRecords)
    return

if __name__ == '__main__':
    main()
